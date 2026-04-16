"""Write the weekly payroll Excel (3 sheets: Summary, Jobs, Parts)."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_CURRENCY = '"$"#,##0.00'
_HEADER_FONT = Font(bold=True)
_TOTAL_FONT = Font(bold=True)
_FINAL_FILL = PatternFill(start_color="FFF4D03F", end_color="FFF4D03F", fill_type="solid")


@dataclass
class WeeklyReportMeta:
    week_start: date
    week_end: date
    run_timestamp: str
    hcp_cache_dir: str
    hcp_response_count: int
    price_sheet_sha8: str
    ticket_url_template: str


@dataclass
class WeeklySummaryRow:
    tech: str
    jobs: int
    gross_revenue: float
    tips: float
    parts_cost: float
    basis: float
    commission: float
    bonuses: float
    overrides: float
    final_pay: float


@dataclass
class JobRow:
    job_number: str
    hcp_id: str
    job_date: date
    customer: str
    amount: float
    tip: float
    parts_cost: float
    basis: float
    listed_techs: str
    owner: str
    primary_comm: float
    charles_bonus: float
    charles_override: float
    notes: str


@dataclass
class PartRow:
    job_number: str
    part_name: str
    quantity: int
    unit_price: float
    total: float
    source: str


@dataclass
class WeeklyReport:
    meta: WeeklyReportMeta
    summary: list[WeeklySummaryRow]
    jobs: list[JobRow]
    parts: list[PartRow]


def _write_meta(ws, meta: WeeklyReportMeta) -> int:
    ws.cell(row=1, column=1, value=f"Run {meta.run_timestamp} | HCP cache: {meta.hcp_cache_dir} | {meta.hcp_response_count} API responses")
    ws.cell(row=2, column=1, value=f"Price sheet SHA: {meta.price_sheet_sha8} | Week: {meta.week_start.isoformat()} to {meta.week_end.isoformat()}")
    return 4


def _write_summary(wb, report: WeeklyReport) -> None:
    ws = wb.create_sheet("Summary")
    start = _write_meta(ws, report.meta)
    headers = ["Tech", "Jobs", "Gross Revenue", "Tips", "Parts Cost", "Basis",
               "Commission", "Bonuses", "Overrides", "Final Pay"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.font = _HEADER_FONT
    for ri, row in enumerate(report.summary, start=start + 1):
        vals = [row.tech, row.jobs, row.gross_revenue, row.tips, row.parts_cost,
                row.basis, row.commission, row.bonuses, row.overrides, row.final_pay]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=ri, column=c, value=v)

    total_row = start + 1 + len(report.summary)
    ws.cell(row=total_row, column=1, value="Total").font = _TOTAL_FONT
    ws.cell(row=total_row, column=2, value=sum(r.jobs for r in report.summary)).font = _TOTAL_FONT
    for col, attr in enumerate(["gross_revenue", "tips", "parts_cost", "basis",
                                 "commission", "bonuses", "overrides", "final_pay"],
                                start=3):
        cell = ws.cell(row=total_row, column=col, value=sum(getattr(r, attr) for r in report.summary))
        cell.font = _TOTAL_FONT

    for ri in range(start + 1, total_row + 1):
        for c in range(3, 11):
            ws.cell(row=ri, column=c).number_format = _CURRENCY
        ws.cell(row=ri, column=10).fill = _FINAL_FILL

    ws.freeze_panes = f"A{start + 1}"
    widths = [22, 6, 15, 10, 12, 14, 14, 11, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_jobs(wb, report: WeeklyReport) -> None:
    ws = wb.create_sheet("Jobs")
    start = _write_meta(ws, report.meta)
    headers = ["Job #", "HCP Id", "Date", "Customer", "Amount", "Tip",
               "Parts Cost", "Basis", "Listed Techs", "Owner",
               "Primary Comm", "Charles Bonus", "Charles Override", "Notes"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.font = _HEADER_FONT

    for ri, j in enumerate(report.jobs, start=start + 1):
        ws.cell(row=ri, column=1, value=j.job_number)
        hcp_cell = ws.cell(row=ri, column=2, value=j.hcp_id)
        if report.meta.ticket_url_template and j.hcp_id:
            hcp_cell.hyperlink = report.meta.ticket_url_template.format(hcp_id=j.hcp_id)
            hcp_cell.font = Font(color="FF0563C1", underline="single")
        ws.cell(row=ri, column=3, value=j.job_date)
        ws.cell(row=ri, column=4, value=j.customer)
        ws.cell(row=ri, column=5, value=j.amount)
        ws.cell(row=ri, column=6, value=j.tip)
        ws.cell(row=ri, column=7, value=j.parts_cost)
        ws.cell(row=ri, column=8, value=j.basis)
        ws.cell(row=ri, column=9, value=j.listed_techs)
        ws.cell(row=ri, column=10, value=j.owner)
        ws.cell(row=ri, column=11, value=j.primary_comm)
        ws.cell(row=ri, column=12, value=j.charles_bonus)
        ws.cell(row=ri, column=13, value=j.charles_override)
        ws.cell(row=ri, column=14, value=j.notes)

    for ri in range(start + 1, start + 1 + len(report.jobs)):
        for c in [5, 6, 7, 8, 11, 12, 13]:
            ws.cell(row=ri, column=c).number_format = _CURRENCY
        ws.cell(row=ri, column=3).number_format = "YYYY-MM-DD"

    ws.freeze_panes = f"A{start + 1}"
    for i, w in enumerate([10, 18, 12, 20, 11, 8, 11, 11, 28, 18, 12, 12, 14, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_parts(wb, report: WeeklyReport) -> None:
    ws = wb.create_sheet("Parts")
    start = _write_meta(ws, report.meta)
    headers = ["Job #", "Part Name", "Qty", "Unit Price", "Total", "Source"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start, column=c, value=h).font = _HEADER_FONT

    for ri, p in enumerate(report.parts, start=start + 1):
        ws.cell(row=ri, column=1, value=p.job_number)
        ws.cell(row=ri, column=2, value=p.part_name)
        ws.cell(row=ri, column=3, value=p.quantity)
        ws.cell(row=ri, column=4, value=p.unit_price)
        ws.cell(row=ri, column=5, value=p.total)
        ws.cell(row=ri, column=6, value=p.source)

    for ri in range(start + 1, start + 1 + len(report.parts)):
        for c in (4, 5):
            ws.cell(row=ri, column=c).number_format = _CURRENCY

    ws.freeze_panes = f"A{start + 1}"
    for i, w in enumerate([10, 32, 6, 12, 12, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_weekly_report(path: Path, report: WeeklyReport) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_summary(wb, report)
    _write_jobs(wb, report)
    _write_parts(wb, report)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
