"""Parse the operator's parts-cost Excel sheet into an in-memory lookup.

Primary table: rows below a header row containing 'Part' and 'Total'.
Supplemental rows: any row below with a non-empty column-A string (or number)
and a positive numeric value somewhere after it. Tolerant by design so the
operator doesn't have to reshape their working file.
"""
from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import openpyxl


@dataclass
class PartPrice:
    name: str
    total: float
    source: Literal["primary", "supplemental"]


class PriceSheet:
    def __init__(self, parts: list[PartPrice]):
        self._parts = {p.name: p for p in parts}
        self._all = parts

    @property
    def part_names(self) -> list[str]:
        return list(self._parts.keys())

    def get(self, name: str) -> PartPrice:
        if name not in self._parts:
            raise KeyError(f"Part not in price sheet: {name!r}")
        return self._parts[name]

    def search(self, query: str, limit: int = 10) -> list[PartPrice]:
        """Fuzzy match: case-insensitive, prefix matches first, then substring."""
        q = query.lower().strip()
        if not q:
            return []
        prefix = [p for p in self._all if p.name.lower().startswith(q)]
        substr = [p for p in self._all
                  if q in p.name.lower() and not p.name.lower().startswith(q)]
        return (prefix + substr)[:limit]

    @staticmethod
    def hash_file(path: Path) -> str:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()


def _find_header_row(ws) -> int:
    for row in range(1, min(ws.max_row + 1, 40)):
        values = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
        str_vals = {str(v).strip() for v in values if v is not None}
        if "Part" in str_vals and "Total" in str_vals:
            return row
    raise ValueError("Could not find price-sheet header row (expected 'Part' and 'Total')")


def _col_index(ws, header_row: int, name: str) -> int:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=header_row, column=c).value or "").strip() == name:
            return c
    raise ValueError(f"Column {name!r} not found in header row {header_row}")


def load_price_sheet(path: Path, *, sheet_name: str = "Pricing") -> PriceSheet:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not in {wb.sheetnames}")
    ws = wb[sheet_name]
    header_row = _find_header_row(ws)
    part_col = _col_index(ws, header_row, "Part")
    total_col = _col_index(ws, header_row, "Total")

    parts: list[PartPrice] = []
    seen: set[str] = set()
    # Primary table: continuous rows with a Part name and numeric Total
    row = header_row + 1
    while row <= ws.max_row:
        name_val = ws.cell(row=row, column=part_col).value
        total_val = ws.cell(row=row, column=total_col).value
        if name_val is None and total_val is None:
            break
        if isinstance(name_val, (int, float)):
            name = str(name_val)
        elif isinstance(name_val, str) and name_val.strip():
            name = name_val.strip()
        else:
            row += 1
            continue
        if isinstance(total_val, (int, float)) and total_val > 0:
            if name not in seen:
                parts.append(PartPrice(name=name, total=float(total_val), source="primary"))
                seen.add(name)
        row += 1

    # Supplemental scan: any later row where col A is a name-like value and
    # there's a positive numeric value in a later column. Take the last numeric.
    for r in range(row, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not vals or vals[0] is None:
            continue
        head = vals[0]
        if isinstance(head, (int, float)):
            name = str(head)
        elif isinstance(head, str) and head.strip():
            name = head.strip()
        else:
            continue
        numerics = [v for v in vals[1:] if isinstance(v, (int, float)) and v > 0]
        if not numerics:
            continue
        total = float(numerics[-1])
        if name not in seen:
            parts.append(PartPrice(name=name, total=total, source="supplemental"))
            seen.add(name)

    return PriceSheet(parts)
