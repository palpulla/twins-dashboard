"""Run once to generate a deterministic price-sheet fixture for tests."""
from pathlib import Path

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pricing"

# Rows 1-7: junk header content (ignored)
ws.cell(row=3, column=4, value="Wayne Dalton")

# Row 8: the real header
headers = ["Part", "List Price", "Multiplier", "Our Price",
           "5.5% Sales Tax", "Energy Surcharge", "Total"]
for col, h in enumerate(headers, start=1):
    ws.cell(row=8, column=col, value=h)

# Rows 9-13: primary parts
primary = [
    (".243 #2 - 30.5\"", 40.95, "0.997", 40.827, 43.072, 1.02, 45.76),
    ("7' Cables", 5.40, "1.109", 5.989, 6.318, 0.150, 8.47),
    ("Drum", 5.73, "1.109", 6.354, 6.704, 0.159, 6.86),
    ("Roller", 1.62, "1.109", 1.796, 1.895, 0.045, 3.50),
    ("Universal Keypad", 38.79, ".82", 43.12, 45.49, 1.08, 47.96),
]
for i, row in enumerate(primary, start=9):
    for col, val in enumerate(row, start=1):
        ws.cell(row=i, column=col, value=val)

# Row 14: blank (end of primary table)
# Row 15: supplemental entry in a different layout (name + trailing number)
ws.cell(row=15, column=1, value=98022)  # numeric "name"
ws.cell(row=15, column=2, value=619.47)
ws.cell(row=15, column=3, value=671.06)

# Row 17: supplemental with a price in the last cell only
ws.cell(row=17, column=1, value="2220L-7")
ws.cell(row=17, column=3, value=316.36)

out = Path(__file__).parent / "parts_sheet_small.xlsx"
wb.save(out)
print(f"wrote {out}")
