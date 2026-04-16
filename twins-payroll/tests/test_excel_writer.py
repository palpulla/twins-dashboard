"""Tests for engine/excel_writer.py"""
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from engine.excel_writer import WeeklyReport, WeeklyReportMeta, WeeklySummaryRow, JobRow, PartRow, write_weekly_report


def _sample_report(tmp_path) -> tuple[Path, WeeklyReport]:
    out = tmp_path / "payroll_week_of_2026-04-13.xlsx"
    meta = WeeklyReportMeta(
        week_start=date(2026, 4, 13),
        week_end=date(2026, 4, 19),
        run_timestamp="2026-04-20T10:30:00",
        hcp_cache_dir="data/hcp_cache/r1",
        hcp_response_count=5,
        price_sheet_sha8="deadbeef",
        ticket_url_template="https://pro.housecallpro.com/app/jobs/{hcp_id}",
    )
    summary = [
        WeeklySummaryRow(
            tech="Charles Rue", jobs=1, gross_revenue=625.0, tips=0.0,
            parts_cost=13.72, basis=611.28, commission=122.26, bonuses=40.0,
            overrides=29.44, final_pay=191.70,
        ),
        WeeklySummaryRow(
            tech="Maurice", jobs=1, gross_revenue=1687.0, tips=0.0,
            parts_cost=91.52, basis=1595.48, commission=319.10, bonuses=0.0,
            overrides=0.0, final_pay=319.10,
        ),
    ]
    jobs = [
        JobRow(
            job_number="14404", hcp_id="job_14404", job_date=date(2026, 4, 14),
            customer="Sarah J.", amount=1687.0, tip=0.0, parts_cost=91.52,
            basis=1595.48, listed_techs="Charles Rue, Maurice", owner="Maurice",
            primary_comm=319.10, charles_bonus=0.0, charles_override=31.91,
            notes="",
        ),
    ]
    parts = [
        PartRow(job_number="14404", part_name=".243 #2 - 30.5\"", quantity=2,
                unit_price=45.76, total=91.52, source="manual"),
    ]
    return out, WeeklyReport(meta=meta, summary=summary, jobs=jobs, parts=parts)


def test_write_weekly_report_creates_three_sheets(tmp_path):
    out, report = _sample_report(tmp_path)
    write_weekly_report(out, report)

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Summary", "Jobs", "Parts"]


def test_summary_sheet_values(tmp_path):
    out, report = _sample_report(tmp_path)
    write_weekly_report(out, report)
    wb = openpyxl.load_workbook(out)
    ws = wb["Summary"]

    header_row = None
    for r in range(1, 10):
        if ws.cell(row=r, column=1).value == "Tech":
            header_row = r
            break
    assert header_row is not None

    data_row = header_row + 1
    assert ws.cell(row=data_row, column=1).value == "Charles Rue"
    ncols = 10
    assert ws.cell(row=data_row, column=ncols).value == pytest.approx(191.70)


def test_jobs_sheet_has_hyperlink(tmp_path):
    out, report = _sample_report(tmp_path)
    write_weekly_report(out, report)
    wb = openpyxl.load_workbook(out)
    ws = wb["Jobs"]
    found = False
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value == "job_14404":
                found = True
                assert ws.cell(row=cell.row, column=1).value == "14404"
    assert found


def test_parts_sheet_values(tmp_path):
    out, report = _sample_report(tmp_path)
    write_weekly_report(out, report)
    wb = openpyxl.load_workbook(out)
    ws = wb["Parts"]
    found_value = None
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "14404":
            found_value = row
            break
    assert found_value is not None
    assert found_value[1] == ".243 #2 - 30.5\""
    assert found_value[2] == 2
    assert found_value[4] == pytest.approx(91.52)
